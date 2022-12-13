import openpyxl
import pickle

arrange_header = ['evaluation_method','threshold','interestWord','top1','top2','top3','top4','top5']

def write_list_in_row(sheet,sheet_i,write_list):
    #一行に書き込み
    for i,value in enumerate(write_list):
        sheet.cell(row=sheet_i[0],column=i+1,value=value)
    sheet_i[0] += 1

def write_items_in_row(word_for_question,sheet,sheet_i,items,evaluation_method,threshold=''):
    for k,v_list in items.items():
        write_list = [evaluation_method,str(threshold)]
        write_list.append(k)
        word_for_question.add(k)
        for v in v_list:
            write_list.append(v['word'])
            word_for_question.add(v['word'])
        write_list_in_row(sheet,sheet_i,write_list)



def arrange_output_and_extract_for_question(output,tweetList):
    #ユーザ名取得
    user_name = output['user_name']
    #アンケート用の単語を入れる変数
    word_for_question = set()
    #キーワードだけ先に入れておく
    for how_to_get_input in ['directly','extractTopic','counter']:
        keywordList = tweetList['keywordsList_'+how_to_get_input]
        for keyword in keywordList:
            word_for_question.add(keyword)

    #Excelにoutputを整形する
    wb = openpyxl.Workbook()
    #wb_sheets = [wb.create_sheet(title=sheetName) for sheetName in ['extractTopic','counter','directly_extractTopic','directly_counter']]

    #各シートにoutputの内容を入れていく
    for sheetName in ['extractTopic','counter','directly_extractTopic','directly_counter']:
        #書きこむoutputを読み込み
        output_sheet = output[sheetName]
        #sheetの作成
        sheet = wb.create_sheet(title=sheetName)
        #Excelは1から（参照渡しで関数内で更新されるようにする）
        sheet_i = [1]
        #ヘッダーの書き込み
        write_list_in_row(sheet,sheet_i,arrange_header)

        #内容の書き込み開始
        for evaluation_method in ['only_cos','wordnet','20_to_20','100_to_20','200_to_100','300_to_300']:
            if evaluation_method == 'only_cos':
                write_items_in_row(word_for_question,sheet,sheet_i,output_sheet[evaluation_method],evaluation_method)
            else:
                for threshold in [(4,7),(3,8),(2,9)]:
                    write_items_in_row(word_for_question,sheet,sheet_i,output_sheet[evaluation_method][threshold],evaluation_method,threshold)


    #シートPathとName
    PathName = './Record/arrange_output_'+user_name+'.xlsx'
    wb.save(PathName)

    #アンケート用のtxtファイルに一応保存
    PathQANameTop = './Record/QAwords_'+user_name
    with open(PathQANameTop+'.txt','w') as fw:
        fw.write('\n'.join(list(word_for_question)))

    #Excel（回答用）のファイル作成
    wbQA = openpyxl.Workbook()
    wsQA = wbQA.active
    wsQA.title = '回答用シート'
    QA_i = [1]
    write_list_in_row(wsQA,QA_i,['Topic','Q1','Q2','Q3'])
    for word in list(word_for_question):
        write_list_in_row(wsQA,QA_i,[word])
    wbQA.save(PathQANameTop+'.xlsx')


if __name__ == '__main__':
    with open('./Record/tweetList_MrSakaikun_10_5.binaryfile','rb') as f:
        tweetList = pickle.load(f)
    with open('./Record/output_MrSakaikun.binaryfile','rb') as f:
        output = pickle.load(f)
    arrange_output_and_extract_for_question(output,tweetList)
